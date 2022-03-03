from pathlib import Path
from download import Download
FILEURL = "https://www.city.kofu.yamanashi.jp/joho/opendata/shisetsu/documents/aedspot-excel.xlsx"
import pandas as pd
from openpyxl import load_workbook
import neologdn
import re

class Yamanashi_AED():
    BASE_DIR = Path(__file__).absolute().parent.parent
    DATA_DIR = BASE_DIR / "data"

    def __init__(self):
        if not self.DATA_DIR.exists():
            self.DATA_DIR.mkdir()
        
        d = Download(FILEURL, self.DATA_DIR)
        d.download()
        self.fname = self.DATA_DIR / d.name
    
    def create_df(self):
        df = pd.read_excel(self.fname, sheet_name="ＡＥＤ設置場所一覧", header=1,usecols="B:E")
        df = df.dropna()
        df.columns = [neologdn.normalize(i) for i in df.columns]
        df = df.applymap(neologdn.normalize)
        self.df = df
    
    def get_data_version(self):
        wb = load_workbook(self.fname)
        sh = wb.worksheets[0]
        return sh.cell(1,5).value

    def query(self, keywords):
        return self.df.loc[self.df["住所"].str.contains(keywords) | self.df["設置施設"].str.contains(keywords)]

if __name__ == "__main__":
    yA = Yamanashi_AED()
    yA.xlsx_to_df()
